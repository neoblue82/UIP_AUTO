import openpyxl
from openpyxl.styles import PatternFill
from datetime import datetime
from datetime import timedelta
import weekMake_var

#오늘 날짜 구해서 파일명
td = datetime.today().strftime("%Y%m%d")

# 현재 스크립트와 같은 폴더에 위치한 엑셀 파일을 읽어옵니다.
wb = openpyxl.load_workbook('D:/rpa/excel/스크래핑.xlsx', read_only=False, data_only=False)

# 엑셀 파일 내 모든 시트 이름을 출력합니다.
# print(wb.sheetnames)

# 활성화된 시트를 새로운 변수에 할당합니다.
ws = wb.active

yellowFill = PatternFill(start_color='fff3bf',
                   end_color='fff3bf',
                   fill_type='solid')

rownum = 0;

ws.column_dimensions["A"].width = '30'
ws.column_dimensions["B"].width = '30'
ws.column_dimensions["C"].width = '80'
ws.column_dimensions["D"].width = '80'

for row in ws.rows:

    삭제문자 = ["안녕하세요", "!!", "melon", "본문", "폰트", "공감하기", "공유하기"
            , "URL복사", " 신고", "폰트 크기", "번역하기", "번역하기", "크게 보기", "이웃 추가"
            , "이웃추가", "기타 기능", "크기 조정", "크기 작게 보기"
            , "하겠습니다.", "같습니다", "입니다"]

    키워드 = row[0].value  # 키워드
    제목 = row[1].value  # 제목
    링크 = row[2].value  # 링크
    내용 = row[3].value  # 내용
    댓글 = row[4].value  # 댓글

    for item in 삭제문자:
        # 문자열 치환
        내용 = 내용.replace(item, "")

    내용수정 = "  ".join(내용.split())
    내용수정 = 내용수정.replace("  ", " ")

    댓글수정 = "  ".join(댓글.split())
    댓글수정 = 댓글수정.replace("  ", " ").replace("답글", "\n").replace("신고", "")
    # 날짜 = row[3].value  # 날짜

    # if(투입공수 is None):
    #     투입공수 = 0
    #
    # print(rownum,"번째 시작=================================================")
    # print(내용수정)

    rownum = rownum + 1
    if(len(내용수정) > 150):
        ws.cell(row=rownum, column=4).value = 내용수정
    else:
        if(rownum != 1):
            ws.cell(row=rownum, column=1).value = '필요없음'
    ws.row_dimensions[rownum].height = '200'

wb.save('D:/rpa/excel/스크래핑_new.xlsx')
print('종료')